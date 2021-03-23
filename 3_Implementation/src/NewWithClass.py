# Author: Sayani Basak(99003772)
# Contact: sayani.basak@ltts.com /basaksayani1997@gmail.com
# Date of creation: 22/3/2021

# -------------------------------------------------------------------------------------------------------------#
# -------------------------------------------------------------------------------------------------------------#

"""
This program performs the task of putting all the data of a particular
candidate or more than one candidate from 5 sheets to a master sheet
provided the user gives the input of name , ps no and email id of that
particular candidate. Finally in the second master sheet it will print
the summary of the first master sheet i.e. it will print the number
of data of the individual person and the total number of data in the
master sheet. This programs uses both pandas and openpyxl library
and it has object oriented programming concepts such as class,
object and function.
"""

# -------------------------------------------------------------------------------------------------------------#
# -------------------------------------------------------------------------------------------------------------#

"""
This program uses pandas and openpyxl library and import pandas as
pd imports the library and rather than using the name pandas,
it's instructed to use the name pd instead. Similarly import
openpyxl as op imports the library and instead of using the
name openpyxl everytime it uses the name op. From Pandas
ExcelWriter is imported in order to write the header in the
first master sheet.
"""
import pandas as pd
from openpyxl import load_workbook

"""
This program uses a class named Read where the workbook function
is defined and has three arguments such as name, ps no and
email id. which has been validated.
"""

#  Global variable
global_count = 1


#  Class
class Read:
    #  Validation Class to validate Name, PS Number, Email
    @staticmethod
    def validation(name1, ps_number, email):
        count = 0  # Local Count
        wb = load_workbook('Data1.xlsx')
        sheets = wb.sheetnames
        for u in range(len(wb.worksheets)):
            sheet = wb[sheets[u]]
            for j in range(2, sheet.max_row + 1):
                if sheet.cell(row=j, column=1).value == ps_number and sheet.cell \
                            (row=j, column=2).value == name1 and sheet.cell \
                            (row=j, column=3).value == email:
                    count += 1
                    break
        if count == 0:
            print("\nData Provided NOT FOUND in DataBase\n")
            globals()['global_count'] = 0
            wb.close()
        else:
            print("\nData Present in Database\n")
            globals()['global_count'] = 1
            wb.close()

    #  This will Read from Excel and Write in Master Sheet and Summary

    @staticmethod
    def readwrite(ps):

        #   WorkBook Load (Sheets Loading in List df)

        sheets = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']
        df = []

        for n in range(0, 5):
            df.append(pd.read_excel(r'Data1.xlsx', sheet_name=sheets[n]))

        # Matching PS Number from all sheets and appending it in a dataframe

        df1 = pd.DataFrame()
        for t in range(0, 5):
            up_d = df[t].loc[(df[t]['Ps No'] == ps)]
            df1 = df1.append(up_d)

        # Merging all data  from sheets into a single row

        d = {'Name': 'first', 'Email': 'first', 'Start Date': 'first',
             'Module Name': 'first', 'Location': 'first', 'Domain': 'first',
             'Duration of Internship': 'first', 'Floor': 'first',
             'Stipend': 'first', 'Gender': 'first', 'Age': 'first',
             'Phone': 'first', 'Education': 'first', 'Profile': 'first',
             'Training Room': 'first', 'Mentor': 'first',
             'Company Name': 'first', 'Address': 'first', 'City': 'first',
             'Country': 'first', 'State': 'first', 'ZIP': 'first',
             'Degree': 'first', 'Semester1': 'first', 'Semester2': 'first',
             'Semester3': 'first', 'Semester4': 'first', 'Semester5': 'first',
             'Semester6': 'first', 'Semester7': 'first', 'Entry Time': 'first',
             'Exit Time': 'first', 'Shift Timings': 'first',
             'Pan Num': 'first', 'Aadhar Num': 'first',
             'Bank Account': 'first', 'End Date': 'first'}

        df1 = df1.groupby('Ps No', as_index=False).aggregate(d) \
            .reindex(columns=df1.columns)

        # Using Openpyxl to create and load dataframe to workbook

        book = load_workbook(r"Data1.xlsx")
        writer = pd.ExcelWriter(r"Data1.xlsx", engine='openpyxl')
        writer.book = book

        # to append data on MasterSheet and not create another Sheet
        writer.sheets = dict(
            (ws.title, ws) for ws in book.worksheets)

        # Create New sheet or Append if MasterSheet exists.

        sheets = book.sheetnames
        if 'MasterSheet' in sheets:
            print("Master Sheet present")
            sheet = book['MasterSheet']
            df1.to_excel(writer, sheet_name='MasterSheet', index=False,
                         header=False, startrow=sheet.max_row)

        # This will create new MasterSheet
        else:
            df1.to_excel(writer, sheet_name='MasterSheet', index=False)

        #  Save the Excel File and Print the Updated sheet
        sheet = book['MasterSheet']
        data = {'Number of Trainers': [sheet.max_row - 1],
                'Individual Data': [sheet.max_column],
                'Total Data': [(sheet.max_row - 1) * sheet.max_column],
                }

        # This will create new Summary

        df2 = pd.DataFrame(data, columns=['Number of Trainers',
                                          'Individual Data', 'Total Data'])
        df2.to_excel(writer, sheet_name='Summary', index=False)

        book.save("Data1.xlsx")
        print(pd.read_excel(r'Data1.xlsx', sheet_name='MasterSheet'))

        # Excel File Close

        book.close()


# d1 = Read()
no_of_inputs = int(input("Select the number of inputs: "))
for i in range(no_of_inputs):
    name = input("Enter the name for Data" + str(i + 1) + " : ")
    try:
        ps_no = int(input("Enter the PS No for Data" + str(i + 1) + " : "))
    except ValueError:
        print("\n!!!!Integer Expected got string!!!!\n")
        continue
    email_id = input("Enter email id for Data" + str(i + 1) + " : ")
    Read.validation(name, ps_no, email_id)
    if global_count == 0:
        continue
    Read.readwrite(ps_no)
